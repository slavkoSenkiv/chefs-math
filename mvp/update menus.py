import openpyxl, os, pprint
from openpyxl.styles import Font
from pathlib import Path
bold = Font(bold=True)
# my notes/don't mind: last letter for iterations is 'c'

# how the menusData structure looks like
"""
menusData = {'menu1': {'menuRecipe1': {'product1': 0, 'product2': 0}, 
                       'menuRecipe2': {'product2': 0, 'product3': 0}},
             'menu2': {'menuRecipe1': {'product1': 0, 'product2': 0},
                       'menuRecipe3': {'product3': 0, 'product4': 0}}}                                           
"""
menusData = {}

for menu in os.listdir(Path.cwd()/'папка меню'):

    menusData.setdefault(menu, {})  # set  / menuData => menu / level in menusData structure

    menuWb = openpyxl.load_workbook(Path.cwd()/'папка меню'/menu)  # load every menu excel workbook one by one and then its 1st sheet where is all the menu data
    menuWbSheet = menuWb.active

    menuWeightPerPerson = 0
    menuCostPerPerson = 0
    menuCaloriesPerPerson = 0

    menuWeightTotal = 0
    menuCostTotal = 0


    # write in menusData db
    # add recipe name level for every menu
    recipeNameIncrement = 1
    for menu_recipes in range(4, menuWbSheet.max_row + 1):
        menuRecipeName = menuWbSheet.cell(row=menu_recipes, column=1).value
        menuRecipeWb = openpyxl.load_workbook(Path.cwd() / 'папка рецептів' / f'{menuRecipeName}.xlsx')
        menuRecipeWbSheet = menuRecipeWb.active
        # next 'if' section is here in case menu has few same recipes
        if menuRecipeName in menusData[menu]:
            newMenuRecipeName = menuRecipeName + str(recipeNameIncrement)
            recipeNameIncrement += 1
            menusData[menu].setdefault(newMenuRecipeName, {})
            # add product level for every recipe in current menu
            for recipeProducts in range(5, menuRecipeWbSheet.max_row + 1):
                # menusData[menu][RecipeName][set productName as value =  recipe item amount * menu recipe portion size * number of guests / recipe output
                menusData[menu][newMenuRecipeName].setdefault(menuRecipeWbSheet.cell(row=recipeProducts, column=1).value, round(menuRecipeWbSheet.cell(row=recipeProducts, column=2).value * menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=1, column=2).value / menuRecipeWbSheet.cell(row=2, column=2).value))
        else:
            menusData[menu].setdefault(menuRecipeName, {})
            # add product level for every recipe in current menu
            for recipeProducts in range(5, menuRecipeWbSheet.max_row + 1):
                # menusData[menu][RecipeName][set productName as value =  recipe item amount * menu recipe portion size * number of guests / recipe output
                menusData[menu][menuRecipeName].setdefault(menuRecipeWbSheet.cell(row=recipeProducts, column=1).value, round(menuRecipeWbSheet.cell(row=recipeProducts, column=2).value * menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=1, column=2).value / menuRecipeWbSheet.cell(row=2, column=2).value))

        # recipe total weight
        recipeWeightOutput = menuRecipeWbSheet.cell(row=2, column=2).value
        # recipe price per 1kg
        menuWbSheet.cell(row=menu_recipes, column=7).value = round(1000 * menuRecipeWbSheet.cell(row=3, column=4).value / recipeWeightOutput, 1)
        # recipe calories per 100
        menuWbSheet.cell(row=menu_recipes, column=8).value = round(100 * menuRecipeWbSheet.cell(row=3, column=5).value / recipeWeightOutput, 1)
        # menu recipe cost per portion
        menuWbSheet.cell(row=menu_recipes, column=3).value = round(menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=menu_recipes, column=7).value / 1000, 1)
        # menu recipe calories per portion
        menuWbSheet.cell(row=menu_recipes, column=4).value = round(menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=menu_recipes, column=8).value / 100, 1)
        # menu recipe weight for all guests
        menuWbSheet.cell(row=menu_recipes, column=5).value = round(menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=1, column=2).value, 1)
        # menu recipe cost for all guests
        menuWbSheet.cell(row=menu_recipes, column=6).value = round(menuWbSheet.cell(row=menu_recipes, column=3).value * menuWbSheet.cell(row=1, column=2).value, 1)

        menuWeightPerPerson += menuWbSheet.cell(row=menu_recipes, column=2).value
        menuCostPerPerson += menuWbSheet.cell(row=menu_recipes, column=3).value
        menuCaloriesPerPerson += menuWbSheet.cell(row=menu_recipes, column=4).value
        menuWeightTotal += menuWbSheet.cell(row=menu_recipes, column=5).value
        menuCostTotal += menuWbSheet.cell(row=menu_recipes, column=6).value

    menuWbSheet.cell(row=2, column=2).value = menuWeightPerPerson
    menuWbSheet.cell(row=2, column=3).value = menuCostPerPerson
    menuWbSheet.cell(row=2, column=4).value = menuCaloriesPerPerson
    menuWbSheet.cell(row=2, column=5).value = menuWeightTotal
    menuWbSheet.cell(row=2, column=6).value = menuCostTotal

    # delete previously created sheets in menu workbook except the 1st one
    firstSheet = menuWb['Sheet1']
    for sheets in menuWb:
        if sheets != firstSheet:
            menuWb.remove(sheets)

    # create sheet shopping list  per recipe
    a = 1
    for menu_recipes in menusData[menu]:
        menuWb.create_sheet(index=a, title=menu_recipes)
        a += 1

        # fulfill the default cells in recipeSheets in menuWb except of Sheet1
        menuWbRecipeShoppingSheet = menuWb[menu_recipes]
        menuWbRecipeShoppingSheet['A1'] = 'назва страви'
        menuWbRecipeShoppingSheet['B1'] = menu_recipes
        menuWbRecipeShoppingSheet['A2'] = 'продукт'
        menuWbRecipeShoppingSheet['B2'] = 'к-ть'
        menuWbRecipeShoppingSheet['A1'].font = bold
        menuWbRecipeShoppingSheet['A2'].font = bold
        menuWbRecipeShoppingSheet['B2'].font = bold

        d = 3
        for menu_recipe_products in menusData[menu][menu_recipes]:
            menuWbRecipeShoppingSheet[f'A{d}'] = menu_recipe_products
            menuWbRecipeShoppingSheet[f'B{d}'] = menusData[menu][menu_recipes][menu_recipe_products]
            d += 1

    menuWb.create_sheet(index=1, title='заг прод')

    menuWb.save(Path.cwd()/'папка меню'/menu)

menusDataDoc = open('menusData.py', 'w', encoding='utf-8')
menusDataDoc.write('menusData = ' + pprint.pformat(menusData))
menusDataDoc.close()






# here starts menu shopping list section
# ______________________________________

import openpyxl, pprint, productsData, menusData
from openpyxl.styles import Font
bold = Font(bold=True)

# how the shopping list data structure looks like
"""
shoppingListData = {'menu1': {'product1': 0, 'product2': 0},
                       'menu2': {'product2': 0, 'product3': 0}}
"""
shoppingListData = {}

# building shopping list per menu in menuData
for menu in menusData.menusData:

    shoppingListData.setdefault(menu, {})  # setting up /shoppingListData => Menu/ level in data structure

    # creating and setting up template for shopping list per menu
    menuWb = openpyxl.load_workbook(Path.cwd()/'папка меню'/menu)

    # delete previously created заг прод sheets in menu workbook
    for sheets in menuWb:
        if sheets.title.startswith('заг прод'):
            menuWb.remove(sheets)

    menuWb.create_sheet(index=1, title='заг прод')
    menuShoppingListSheet = menuWb['заг прод']

    menuShoppingListSheet['A1'] = 'назва меню'
    menuShoppingListSheet['B1'] = 'заг кть'
    menuShoppingListSheet['C1'] = 'заг варт'
    menuShoppingListSheet['A3'] = 'продукт'
    menuShoppingListSheet['B3'] = 'к-ть'
    menuShoppingListSheet['C3'] = 'вартість'
    menuShoppingListSheet['D3'] = 'ціна'
    menuShoppingListSheet['E3'] = '% вартості'
    menuShoppingListSheet['F3'] = '% ціни'
    menuShoppingListSheet['A2'] = menu[:-5]
    menuShoppingListSheet['A1'].font = bold
    menuShoppingListSheet['B1'].font = bold
    menuShoppingListSheet['C1'].font = bold
    menuShoppingListSheet['A3'].font = bold
    menuShoppingListSheet['B3'].font = bold
    menuShoppingListSheet['C3'].font = bold
    menuShoppingListSheet['D3'].font = bold
    menuShoppingListSheet['E3'].font = bold
    menuShoppingListSheet['F3'].font = bold

    for menuRecipes in menusData.menusData[menu]:
        for product in menusData.menusData[menu][menuRecipes]:
            shoppingListData[menu].setdefault(product, 0)
            shoppingListData[menu][product] += menusData.menusData[menu][menuRecipes][product]

    menuWb.save(Path.cwd()/'папка меню'/menu)

    # upper everything is ok _________________________________-

for menus in shoppingListData:
    menuShoppingListWb = openpyxl.load_workbook(Path.cwd()/'папка меню'/menus)
    menuShoppingListWbSheet = menuShoppingListWb['заг прод']

    totalShoppingListWeight = 0
    totalShoppingListCost = 0
    totalShoppingListPrice = 0

    for products in shoppingListData[menus]:
        # product name in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row + 1, column=1).value = products

        # product weight in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=2).value = shoppingListData[menus][products]
        totalShoppingListWeight += menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=2).value

        # product price in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=4).value = productsData.productsData[products]['price']
        totalShoppingListPrice += menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=4).value

        # product cost in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=3).value = round(menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=2).value * productsData.productsData[products]['price'] / 1000, 1)
        totalShoppingListCost += menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=3).value

    menuShoppingListWbSheet['B2'] = totalShoppingListWeight
    menuShoppingListWbSheet['C2'] = totalShoppingListCost
    menuShoppingListWbSheet['D2'] = totalShoppingListPrice

    for products in range(4, menuShoppingListWbSheet.max_row + 1):
        menuShoppingListWbSheet.cell(row=products, column=5).value = round(menuShoppingListWbSheet.cell(row=products, column=3).value * 100 / totalShoppingListCost)
        menuShoppingListWbSheet.cell(row=products, column=6).value = round(menuShoppingListWbSheet.cell(row=products, column=4).value * 100 / totalShoppingListPrice)

    menuShoppingListWb.save(Path.cwd()/'папка меню'/menus)

shoppingListDoc = open('shoppingListData.py', 'w', encoding='utf-8')
shoppingListDoc.write('shoppingListData = ' + pprint.pformat(shoppingListData))
shoppingListDoc.close()







