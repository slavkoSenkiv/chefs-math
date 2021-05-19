import openpyxl, os

recipesListWb = openpyxl.load_workbook('список рецептів.xlsx')
recipesListSheet = recipesListWb.active

recipesListRow = 2
for recipe in os.listdir():
    if recipe.startswith('рецепт'):
        recipeWb = openpyxl.load_workbook(recipe)
        recipeSheet = recipeWb.active

        recipeName = recipe[:-5]
        recipeWeight = recipeSheet.cell(row=2, column=2).value
        recipeCost = round(recipeSheet.cell(row=3, column=4).value, 1)
        recipeCaloriesReal = round(recipeSheet.cell(row=3, column=5).value)
        recipePrice100 = round(recipeCost / recipeWeight * 100, 1)
        recipeCalories100 = round(recipeCaloriesReal / recipeWeight * 100, 1)

        recipesListSheet.cell(row=recipesListRow, column=1).value = recipeName
        recipesListSheet.cell(row=recipesListRow, column=2).value = recipeWeight
        recipesListSheet.cell(row=recipesListRow, column=3).value = recipeCost
        recipesListSheet.cell(row=recipesListRow, column=4).value = recipeCaloriesReal
        recipesListSheet.cell(row=recipesListRow, column=5).value = recipePrice100
        recipesListSheet.cell(row=recipesListRow, column=6).value = recipeCalories100
        recipesListRow += 1

recipesListWb.save('список рецептів.xlsx')
