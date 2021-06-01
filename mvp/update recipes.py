import openpyxl, os
from pathlib import Path
import pprint

# update productsData
productsData = {}
productWb = openpyxl.load_workbook('список продуктів.xlsx')
productSheet = productWb.active
""" 
# how productsData structure looks like
productsData = {'product_name': 
                    {'product_unit': str, 
                     'product_unit_weight': int, 
                     'product_price': int, 
                     'product_calories': int, 
                     'product_netWeightPercent': int}}
"""

for products in range(2, productSheet.max_row + 1):
    product_name = productSheet['A' + str(products)].value
    product_price = productSheet['B' + str(products)].value
    product_calories = productSheet['C' + str(products)].value
    product_netWeightPercent = productSheet['D' + str(products)].value

    # make sure the key for this product exists
    productsData.setdefault(product_name, {'price': 0, 'calories': 0, 'netWeightPercent': 0})

    productsData[product_name]['price'] = product_price
    productsData[product_name]['calories'] = product_calories
    productsData[product_name]['netWeightPercent'] = product_netWeightPercent

for recipe in os.listdir(Path.cwd()/'папка рецептів'):


    recipeWb = openpyxl.load_workbook(Path.cwd()/'папка рецептів'/recipe)
    recipeSheet = recipeWb.active

    total_recipe_gross_weight = 0
    total_recipe_net_weight = 0
    total_recipe_cost = 0
    total_recipe_caloriesReal = 0

    for recipe_product in range(5, recipeSheet.max_row + 1):
        recipe_product_gross_weight = recipeSheet.cell(row=recipe_product, column=2).value
        recipe_product_net_weight = recipe_product_gross_weight * (productsData[recipeSheet.cell(row=recipe_product, column=1).value]['netWeightPercent']) / 100
        # Expl: net_weight = gross_weight * % net / 100
        recipe_product_price = productsData[recipeSheet.cell(row=recipe_product, column=1).value]['price']
        recipe_product_calories100 = productsData[recipeSheet.cell(row=recipe_product, column=1).value]['calories']
        recipe_product_cost  = round(recipe_product_price * recipe_product_gross_weight / 1000, 1)
        recipe_product_caloriesReal  = round(recipe_product_calories100 * recipe_product_net_weight / 100)

        recipeSheet.cell(row=recipe_product, column=3).value = recipe_product_net_weight
        recipeSheet.cell(row=recipe_product, column=4).value = recipe_product_cost
        recipeSheet.cell(row=recipe_product, column=5).value = recipe_product_caloriesReal
        recipeSheet.cell(row=recipe_product, column=6).value = recipe_product_price
        recipeSheet.cell(row=recipe_product, column=7).value = recipe_product_calories100

        # total_recipe_gross_weight += recipeSheet.cell(row=recipe_product, column=2).value
        total_recipe_gross_weight += recipe_product_gross_weight
        total_recipe_net_weight += recipe_product_net_weight
        total_recipe_cost += recipe_product_cost
        total_recipe_caloriesReal += recipe_product_caloriesReal

    recipeSheet.cell(row=3, column=2).value = total_recipe_gross_weight
    recipeSheet.cell(row=3, column=3).value = total_recipe_net_weight
    recipeSheet.cell(row=3, column=4).value = total_recipe_cost
    recipeSheet.cell(row=3, column=5).value = total_recipe_caloriesReal

    for recipe_product in range(5, recipeSheet.max_row + 1):
        recipeSheet.cell(row=recipe_product, column=8).value = round(recipeSheet.cell(row=recipe_product, column=3).value * 100 / total_recipe_net_weight)
        recipeSheet.cell(row=recipe_product, column=9).value = round(recipeSheet.cell(row=recipe_product, column=4).value * 100 / total_recipe_cost)
        recipeSheet.cell(row=recipe_product, column=10).value = round(recipeSheet.cell(row=recipe_product, column=5).value * 100 / total_recipe_caloriesReal)

    recipeWb.save(Path.cwd()/'папка рецептів'/recipe)

productsDataDoc = open('productsData.py', 'w', encoding='utf-8')
productsDataDoc.write('productsData = ' + pprint.pformat(productsData))
productsDataDoc.close()




# update recipe list part of the script
# _____________________________________

recipesListWb = openpyxl.load_workbook('список рецептів.xlsx')
recipesListSheet = recipesListWb.active

recipesListRow = 2
for recipe in os.listdir(Path.cwd()/'папка рецептів'):

    recipeWb = openpyxl.load_workbook(Path.cwd()/'папка рецептів'/recipe)
    recipeSheet = recipeWb.active

    recipeName = recipe[:-5]
    recipeWeight = recipeSheet.cell(row=2, column=2).value
    recipeCost = round(recipeSheet.cell(row=3, column=4).value, 1)
    recipeCaloriesReal = round(recipeSheet.cell(row=3, column=5).value)
    recipePrice100 = round(recipeCost / recipeWeight * 100, 1)
    recipeCalories100 = round(recipeCaloriesReal / recipeWeight * 100, 1)

    recipesListSheet.cell(row=recipesListRow, column=1).value = recipeName
    recipesListSheet.cell(row=recipesListRow, column=2).value = recipeWeight
    recipesListSheet.cell(row=recipesListRow, column=3).value = recipeCost
    recipesListSheet.cell(row=recipesListRow, column=4).value = recipeCaloriesReal
    recipesListSheet.cell(row=recipesListRow, column=5).value = recipePrice100
    recipesListSheet.cell(row=recipesListRow, column=6).value = recipeCalories100
    recipesListRow += 1

recipesListWb.save('список рецептів.xlsx')
