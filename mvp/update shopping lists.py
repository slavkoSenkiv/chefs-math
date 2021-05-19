import openpyxl, pprint, productsData, menusData
from openpyxl.styles import Font
bold = Font(bold=True)

# how the shopping list data structure looks like
"""
shoppingListData = {'menu1': {'product1': 0, 'product2': 0},
                       'menu2': {'product2': 0, 'product3': 0}}
"""
shoppingListData = {}

# building shopping list per menu in menuData
for menu in menusData.menusData:

    shoppingListData.setdefault(menu, {})  # setting up /shoppingListData => Menu/ level in data structure
    print('menu', menu)

    # creating and setting up template for shopping list per menu
    menuShoppingList = openpyxl.Workbook()
    menuShoppingListSheet = menuShoppingList.active

    menuShoppingListSheet['A1'] = 'назва меню'
    menuShoppingListSheet['B1'] = 'заг кть'
    menuShoppingListSheet['C1'] = 'заг варт'
    menuShoppingListSheet['A3'] = 'продукт'
    menuShoppingListSheet['B3'] = 'к-ть'
    menuShoppingListSheet['C3'] = 'вартість'
    menuShoppingListSheet['D3'] = 'ціна'
    menuShoppingListSheet['E3'] = '% вартості'
    menuShoppingListSheet['F3'] = '% ціни'
    menuShoppingListSheet['A2'] = menu[:-5]
    menuShoppingListSheet['A1'].font = bold
    menuShoppingListSheet['B1'].font = bold
    menuShoppingListSheet['C1'].font = bold
    menuShoppingListSheet['A3'].font = bold
    menuShoppingListSheet['B3'].font = bold
    menuShoppingListSheet['C3'].font = bold
    menuShoppingListSheet['D3'].font = bold
    menuShoppingListSheet['E3'].font = bold
    menuShoppingListSheet['F3'].font = bold

    for menuRecipes in menusData.menusData[menu]:
        print('     menuRecipes', menuRecipes)
        for product in menusData.menusData[menu][menuRecipes]:
            print('         product', product)
            shoppingListData[menu].setdefault(product, 0)
            shoppingListData[menu][product] += menusData.menusData[menu][menuRecipes][product]

    menuShoppingList.save('шопінг ' + menu)

    # upper everything is ok _________________________________-

for menus in shoppingListData:
    menuShoppingListWb = openpyxl.load_workbook('шопінг ' + menus)
    menuShoppingListWbSheet = menuShoppingListWb.active

    totalShoppingListWeight = 0
    totalShoppingListCost = 0
    totalShoppingListPrice = 0

    for products in shoppingListData[menus]:
        # product name in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row + 1, column=1).value = products

        # product weight in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=2).value = shoppingListData[menus][products]
        totalShoppingListWeight += menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=2).value

        # product price in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=4).value = productsData.productsData[products]['price']
        totalShoppingListPrice += menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=4).value

        # product cost in shopping list
        menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=3).value = round(menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=2).value * productsData.productsData[products]['price'] / 1000, 1)
        totalShoppingListCost += menuShoppingListWbSheet.cell(row=menuShoppingListWbSheet.max_row, column=3).value

    menuShoppingListWbSheet['B2'] = totalShoppingListWeight
    menuShoppingListWbSheet['C2'] = totalShoppingListCost
    menuShoppingListWbSheet['D2'] = totalShoppingListPrice

    for products in range(4, menuShoppingListWbSheet.max_row + 1):
        menuShoppingListWbSheet.cell(row=products, column=5).value = round(menuShoppingListWbSheet.cell(row=products, column=3).value * 100 / totalShoppingListCost)
        menuShoppingListWbSheet.cell(row=products, column=6).value = round(menuShoppingListWbSheet.cell(row=products, column=4).value * 100 / totalShoppingListPrice)

    menuShoppingListWb.save('шопінг ' + menus)

shoppingListDoc = open('shoppingListData.py', 'w', encoding='utf-8')
shoppingListDoc.write('shoppingListData = ' + pprint.pformat(shoppingListData))
shoppingListDoc.close()
