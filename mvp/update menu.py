import openpyxl
from pathlib import Path
import pprint

menuWb = openpyxl.load_workbook('меню_1.xlsx')
menuSheet = menuWb.active
# menuProductsData = {'картопля': 2, 'капуста': 3, ...}}
menuProductsData = {}

# load one by one menu recipes sheets
for menu_recipes in range(9, menuSheet.max_row + 1):
    recipe_name = f'{menuSheet.cell(column=1, row=menu_recipes).value}.xlsx'
    recipeWb = openpyxl.load_workbook(recipe_name)
    recipeSheet = recipeWb.active

    # load and add recipes items to menuData
    for recipe_product in range(4, recipeSheet.max_row + 1):
        menuProductsData.setdefault(recipeSheet.cell(column=1, row=recipe_product).value)
    for

print(pprint.pformat(menuProductsData))

