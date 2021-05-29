import openpyxl, os, pprint
from openpyxl.styles import Font
bold = Font(bold=True)
# my notes/don't mind: last letter for iterations is 'c'

# how the menusData structure looks like
"""
menusData = {'menu1': {'menuRecipe1': {'product1': 0, 'product2': 0}, 
                       'menuRecipe2': {'product2': 0, 'product3': 0}},
             'menu2': {'menuRecipe1': {'product1': 0, 'product2': 0},
                       'menuRecipe3': {'product3': 0, 'product4': 0}}}                                           
"""
menusData = {}

for menu in os.listdir():
    if menu.startswith('меню'):

        menusData.setdefault(menu, {})  # set  / menuData => menu / level in menusData structure

        menuWb = openpyxl.load_workbook(menu)  # load every menu excel workbook one by one and then its 1st sheet where is all the menu data
        menuWbSheet = menuWb.active

        menuWeightPerPerson = 0
        menuCostPerPerson = 0
        menuCaloriesPerPerson = 0

        menuWeightTotal = 0
        menuCostTotal = 0

        # add recipe name level for current menu that loop is working with to menusData db
        recipeNameIncrement = 1
        for menu_recipes in range(4, menuWbSheet.max_row + 1):
            menuRecipeName = menuWbSheet.cell(row=menu_recipes, column=1).value
            # next 'if' section is here in case menu has few same recipes
            if menuRecipeName in menusData[menu]:
                menusData[menu].setdefault(menuRecipeName + str(recipeNameIncrement), {})
                recipeNameIncrement += 1
            else:
                menusData[menu].setdefault(menuRecipeName, {})

            menuRecipeWb = openpyxl.load_workbook(menuRecipeName + '.xlsx')
            menuRecipeWbSheet = menuRecipeWb.active

            # write in menusData db
            # add next level to menusData[menu][menuRecipeName] => menuRecipeProducts as a keys and (these products weight * portions size * guest number / 1000) as values
            for recipeProducts in range(5, menuRecipeWbSheet.max_row + 1):
                # menusData[menu][menuRecipeName].setdefault(recipe_product as key, recipe_gross_weight * menu_recipe_portion_size * guest_number / 1000
                if menuRecipeName in menusData[menu]:
                    menusData[menu][menuRecipeName].setdefault(menuRecipeWbSheet.cell(row=recipeProducts, column=1).value, menuRecipeWbSheet.cell(row=recipeProducts, column=2).value * menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=1, column=2).value / 1000)
                else:
                    menusData[menu][menuRecipeName + str(c)].setdefault(menuRecipeWbSheet.cell(row=recipeProducts, column=1).value, menuRecipeWbSheet.cell(row=recipeProducts, column=2).value * menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=1, column=2).value / 1000)


            # recipe total weight
            recipeWeightOutput = menuRecipeWbSheet.cell(row=2, column=2).value
            # recipe price per 1kg
            menuWbSheet.cell(row=menu_recipes, column=7).value = round(1000 * menuRecipeWbSheet.cell(row=3, column=4).value / recipeWeightOutput, 1)
            # recipe calories per 100
            menuWbSheet.cell(row=menu_recipes, column=8).value = round(100 * menuRecipeWbSheet.cell(row=3, column=5).value / recipeWeightOutput, 1)
            # menu recipe cost per portion
            menuWbSheet.cell(row=menu_recipes, column=3).value = round(menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=menu_recipes, column=7).value / 1000, 1)
            # menu recipe calories per portion
            menuWbSheet.cell(row=menu_recipes, column=4).value = round(menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=menu_recipes, column=8).value / 100, 1)
            # menu recipe weight for all guests
            menuWbSheet.cell(row=menu_recipes, column=5).value = round(menuWbSheet.cell(row=menu_recipes, column=2).value * menuWbSheet.cell(row=1, column=2).value, 1)
            # menu recipe cost for all guests
            menuWbSheet.cell(row=menu_recipes, column=6).value = round(menuWbSheet.cell(row=menu_recipes, column=3).value * menuWbSheet.cell(row=1, column=2).value, 1)

            menuWeightPerPerson += menuWbSheet.cell(row=menu_recipes, column=2).value
            menuCostPerPerson += menuWbSheet.cell(row=menu_recipes, column=3).value
            menuCaloriesPerPerson += menuWbSheet.cell(row=menu_recipes, column=4).value
            menuWeightTotal += menuWbSheet.cell(row=menu_recipes, column=5).value
            menuCostTotal += menuWbSheet.cell(row=menu_recipes, column=6).value

        menuWbSheet.cell(row=2, column=2).value = menuWeightPerPerson
        menuWbSheet.cell(row=2, column=3).value = menuCostPerPerson
        menuWbSheet.cell(row=2, column=4).value = menuCaloriesPerPerson
        menuWbSheet.cell(row=2, column=5).value = menuWeightTotal
        menuWbSheet.cell(row=2, column=6).value = menuCostTotal

        # delete previously created sheets in menu workbook except the 1st one
        firstSheet = menuWb['Sheet1']
        for sheets in menuWb:
            if sheets != firstSheet:
                menuWb.remove(sheets)

        # create sheet shopping list  per recipe
        a = 1
        for menu_recipes in menusData[menu]:
            menuWb.create_sheet(index=a, title=menu_recipes[7:])
            a += 1

            # fulfill the default cells in recipeSheets in menuWb except of Sheet1
            menuWbRecipeShoppingSheet = menuWb[menu_recipes[7:]]
            menuWbRecipeShoppingSheet['A1'] = 'назва страви'
            menuWbRecipeShoppingSheet['B1'] = menu_recipes[7:]
            menuWbRecipeShoppingSheet['A2'] = 'продукт'
            menuWbRecipeShoppingSheet['B2'] = 'к-ть'
            menuWbRecipeShoppingSheet['A1'].font = bold
            menuWbRecipeShoppingSheet['A2'].font = bold
            menuWbRecipeShoppingSheet['B2'].font = bold

            d = 3
            for menu_recipe_products in menusData[menu][menu_recipes]:
                menuWbRecipeShoppingSheet[f'A{d}'] = menu_recipe_products
                menuWbRecipeShoppingSheet[f'B{d}'] = menusData[menu][menu_recipes][menu_recipe_products]
                d += 1

        menuWb.save(menu)

menusDataDoc = open('menusData.py', 'w', encoding='utf-8')
menusDataDoc.write('menusData = ' + pprint.pformat(menusData))
menusDataDoc.close()






